"""
DART 재무분석기 v9 — FastAPI 백엔드
v8 account_id 기반 매핑 + 3개 회사 멀티비교 + Claude CLI AI 분석
포트: 8002
"""
import io, zipfile, os, json, datetime, subprocess, shutil
from pathlib import Path
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
import urllib.request, xml.etree.ElementTree as ET

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("pip install openpyxl")

# ── 설정 ─────────────────────────────────────────────
DART_API_KEY = "f8692c12c4fad4928eabbfe55e957a4d29fef157"
BASE_DIR     = Path(__file__).resolve().parent
# Vercel 서버리스: /tmp 만 쓰기 가능
TMP_DIR      = Path("/tmp") if Path("/tmp").exists() else BASE_DIR
CACHE_XML    = TMP_DIR / "dart_corpcode.xml"
OUTPUT_DIR   = TMP_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)
YEARS        = [2023, 2024, 2025]

# ── account_id → 표준항목 (v8 동일) ──────────────────
CODE_MAP = {
    "ifrs-full_Revenue":                                        "매출",
    "ifrs-full_RevenueFromContractsWithCustomers":              "매출",
    "ifrs_Revenue":                                             "매출",
    "dart_Revenue":                                             "매출",
    "dart_OperatingRevenue":                                    "매출",
    "ifrs-full_CostOfSales":                                    "매출원가",
    "ifrs_CostOfSales":                                         "매출원가",
    "ifrs-full_GrossProfit":                                    "매출총이익",
    "ifrs_GrossProfit":                                         "매출총이익",
    "ifrs-full_SellingGeneralAndAdministrativeExpense":         "판매비와관리비",
    "dart_TotalSellingGeneralAdministrativeExpenses":           "판매비와관리비",
    "ifrs-full_DistributionCosts":                              "판매비와관리비",
    "ifrs-full_ProfitLossFromOperatingActivities":              "영업이익",
    "dart_OperatingIncomeLoss":                                 "영업이익",
    "ifrs_ProfitLossFromOperatingActivities":                   "영업이익",
    "ifrs-full_FinanceIncome":                                  "금융수익",
    "ifrs_FinanceIncome":                                       "금융수익",
    "ifrs-full_FinanceCosts":                                   "금융비용",
    "ifrs_FinanceCosts":                                        "금융비용",
    "ifrs-full_ProfitLossBeforeTax":                            "법인세비용차감전순이익",
    "ifrs_ProfitLossBeforeTax":                                 "법인세비용차감전순이익",
    "ifrs-full_IncomeTaxExpenseContinuingOperations":           "법인세비용",
    "ifrs_IncomeTaxExpenseContinuingOperations":                "법인세비용",
    "ifrs-full_ProfitLoss":                                     "당기순이익",
    "ifrs_ProfitLoss":                                          "당기순이익",
    "ifrs-full_BasicEarningsLossPerShare":                      "EPS",
    "ifrs_BasicEarningsLossPerShare":                           "EPS",
    "ifrs-full_BasicEarningsPerShare":                          "EPS",
    "ifrs-full_CurrentAssets":                                  "유동자산",
    "ifrs_CurrentAssets":                                       "유동자산",
    "ifrs-full_NoncurrentAssets":                               "비유동자산",
    "ifrs_NoncurrentAssets":                                    "비유동자산",
    "ifrs-full_Assets":                                         "자산총계",
    "ifrs_Assets":                                              "자산총계",
    "ifrs-full_CurrentLiabilities":                             "유동부채",
    "ifrs_CurrentLiabilities":                                  "유동부채",
    "ifrs-full_NoncurrentLiabilities":                          "비유동부채",
    "ifrs_NoncurrentLiabilities":                               "비유동부채",
    "ifrs-full_Liabilities":                                    "부채총계",
    "ifrs_Liabilities":                                         "부채총계",
    "ifrs-full_RetainedEarnings":                               "이익잉여금",
    "ifrs_RetainedEarnings":                                    "이익잉여금",
    "ifrs-full_Equity":                                         "자본총계",
    "ifrs_Equity":                                              "자본총계",
    "ifrs-full_EquityAttributableToOwnersOfParent":             "자본총계",
}

NAME_MAP = {
    "보통주기본주당이익(손실)":           "EPS",
    "기본주당순이익":                     "EPS",
    "보통주기본주당순이익(손실)":         "EPS",
    "보통주 기본 및 희석주당이익":        "EPS",
    "기본주당계속영업이익(손실)":         "EPS",
    "보통주기본주당계속영업이익(손실)":   "EPS",
    "매출액":                             "매출",
    "영업수익":                           "매출",
    "수익(매출액)":                       "매출",
    "도급공사수익":                       "매출",
    "영업이익(손실)":                     "영업이익",
    "영업이익(손실)의계":                 "영업이익",
    "매출총이익(손실)":                   "매출총이익",
    "금융원가":                           "금융비용",
    "법인세비용차감전순이익(손실)":       "법인세비용차감전순이익",
    "법인세비용차감전순손실":             "법인세비용차감전순이익",
    "법인세비용(수익)":                   "법인세비용",
    "당기순이익(손실)":                   "당기순이익",
    "당기순손익":                         "당기순이익",
    "이익잉여금(결손금)":                 "이익잉여금",
}

# ── Claude CLI ───────────────────────────────────────
def find_claude_cli():
    if shutil.which("claude"): return shutil.which("claude")
    for p in [
        Path(os.environ.get("APPDATA","")) / "npm" / "claude.cmd",
        Path(r"C:\Users\khcho\AppData\Roaming\npm\claude.cmd"),
    ]:
        if p.exists(): return str(p)
    return None

CLAUDE_CLI = find_claude_cli()

# ── FastAPI ───────────────────────────────────────────
app = FastAPI(title="DART 재무분석기 v9")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ── 기업코드 XML ──────────────────────────────────────
def get_corp_xml():
    if CACHE_XML.exists():
        age = datetime.datetime.now() - datetime.datetime.fromtimestamp(CACHE_XML.stat().st_mtime)
        if age.days < 7:
            return ET.parse(CACHE_XML).getroot()
    with urllib.request.urlopen(
        f"https://opendart.fss.or.kr/api/corpCode.xml?crtfc_key={DART_API_KEY}", timeout=30
    ) as r:
        zdata = r.read()
    with zipfile.ZipFile(io.BytesIO(zdata)) as z:
        xml_bytes = z.read(z.namelist()[0])
    CACHE_XML.write_bytes(xml_bytes)
    return ET.fromstring(xml_bytes)

# ── DART fetch (account_id 기반) ──────────────────────
def fetch_by_code(corp_code: str, year: int) -> dict:
    for fs in ("CFS", "OFS"):
        url = (f"https://opendart.fss.or.kr/api/fnlttSinglAcntAll.json"
               f"?crtfc_key={DART_API_KEY}&corp_code={corp_code}"
               f"&bsns_year={year}&reprt_code=11011&fs_div={fs}")
        try:
            with urllib.request.urlopen(url, timeout=30) as r:
                d = json.load(r)
        except: continue
        if d.get("status") != "000" or not d.get("list"): continue
        result = {}
        for it in d["list"]:
            code = it.get("account_id","").strip()
            std  = CODE_MAP.get(code)
            if not std:
                std = NAME_MAP.get(it.get("account_nm","").strip())
            if std and std not in result:
                val = it.get("thstrm_amount","")
                try:    result[std] = int(val.replace(",",""))
                except: result[std] = None
        result["_fs"] = fs
        return result
    return {}

# ── 투자지표 계산 ─────────────────────────────────────
def add_ratios(d: dict) -> dict:
    def pct(a, b):
        try: return round(a/b*100, 2) if a is not None and b else None
        except: return None
    def rat(a, b):
        try: return round(a/b, 2) if a is not None and b else None
        except: return None
    d["영업이익률"]   = pct(d.get("영업이익"),   d.get("매출"))
    d["순이익률"]     = pct(d.get("당기순이익"), d.get("매출"))
    d["ROE"]          = pct(d.get("당기순이익"), d.get("자본총계"))
    d["ROA"]          = pct(d.get("당기순이익"), d.get("자산총계"))
    d["부채비율"]     = pct(d.get("부채총계"),   d.get("자본총계"))
    d["유동비율"]     = pct(d.get("유동자산"),   d.get("유동부채"))
    d["이자보상배율"] = rat(d.get("영업이익"),   d.get("금융비용"))
    return d

# ── Claude CLI AI 의견 ────────────────────────────────
def get_ai_opinion(corp_name: str, corp_data: dict) -> list[str]:
    if not CLAUDE_CLI:
        return ["Claude CLI를 찾을 수 없습니다."]
    prompt = "\n".join([
        f"당신은 대한민국 최고의 기업 재무분석 전문가입니다.",
        f"아래 {corp_name}의 DART 재무 데이터를 분석하여",
        f"①~⑩ 번호로 시작하는 정확히 10줄 한국어 의견을 작성하세요.",
        f"각 줄에 구체적 수치를 반드시 포함하고, ① 부터 바로 시작하세요.",
        f"",
        f"=== {corp_name} 재무 데이터 ===",
        json.dumps(corp_data, ensure_ascii=False, indent=2),
        f"",
        "분석: ① 매출성장성 ② 영업이익률 ③ 순이익추세 ④ ROE ⑤ ROA ⑥ 부채비율 ⑦ 유동비율 ⑧ 이자보상배율 ⑨ 3개년추세 ⑩ 종합평가",
        "① 부터 ⑩ 까지 정확히 10줄만 출력.",
    ])
    try:
        r = subprocess.run("claude --print -", input=prompt, capture_output=True,
                           text=True, encoding="utf-8", timeout=120, shell=True)
        if r.returncode != 0: return [f"Claude CLI 오류 (코드 {r.returncode})"]
        numbered = ["①","②","③","④","⑤","⑥","⑦","⑧","⑨","⑩"]
        lines = [l.strip() for l in r.stdout.splitlines()
                 if l.strip() and any(l.strip().startswith(n) for n in numbered)]
        return lines[:10] if lines else [r.stdout.strip()[:500]]
    except Exception as e:
        return [f"오류: {e}"]

# ── API 엔드포인트 ────────────────────────────────────
class SearchReq(BaseModel):
    name: str

class MultiAnalyzeReq(BaseModel):
    corps: list[dict]   # [{"corp_code":"...","corp_name":"..."}] × 3

class OpinionReq(BaseModel):
    corp_name: str
    data: dict          # 연도별 재무 데이터

@app.post("/api/search")
def search(req: SearchReq):
    try: root = get_corp_xml()
    except Exception as e: raise HTTPException(500, str(e))
    q = req.name.strip()
    results = []
    for corp in root.findall("list"):
        nm = (corp.findtext("corp_name") or "").strip()
        if q in nm:
            results.append({
                "corp_code":  corp.findtext("corp_code","").strip(),
                "corp_name":  nm,
                "stock_code": corp.findtext("stock_code","").strip(),
            })
    results.sort(key=lambda x: (x["stock_code"]=="", len(x["corp_name"])))
    return {"results": results[:30]}

@app.post("/api/analyze-multi")
def analyze_multi(req: MultiAnalyzeReq):
    if len(req.corps) != 3:
        raise HTTPException(400, "3개 기업을 입력하세요.")
    result_corps = []
    for corp in req.corps:
        yr_data = {}
        for yr in YEARS:
            raw = fetch_by_code(corp["corp_code"], yr)
            yr_data[yr] = add_ratios(raw)
        result_corps.append({
            "corp_name": corp["corp_name"],
            "data": {str(yr): {k:v for k,v in yr_data[yr].items() if not k.startswith("_")}
                     for yr in YEARS},
        })
    # Excel 생성
    today = datetime.date.today().strftime("%Y%m%d")
    names = "_".join(c["corp_name"][:4] for c in req.corps)
    filename = f"멀티비교_{names}_{today}.xlsx"
    _make_excel(result_corps, OUTPUT_DIR / filename)
    return {"corps": result_corps, "filename": filename}

@app.post("/api/opinion")
def opinion(req: OpinionReq):
    lines = get_ai_opinion(req.corp_name, req.data)
    return {"corp_name": req.corp_name, "lines": lines}

@app.get("/api/download/{filename}")
def download(filename: str):
    safe  = Path(filename).name
    fpath = OUTPUT_DIR / safe
    if not fpath.exists(): raise HTTPException(404, "파일 없음")
    return FileResponse(path=fpath, filename=safe,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/")
def root(): return {"message": "DART v9 API — /docs"}

# ── Excel 생성 ────────────────────────────────────────
def _make_excel(corps: list, path: Path):
    wb = Workbook()
    thin = Side(style="thin", color="D0D0D0")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    def fill(c): return PatternFill("solid", fgColor=c)
    def fa(v):
        if v is None: return "-"
        return f"{'-' if v<0 else ''}{abs(v)//100_000_000:,}억"
    def fp(v): return f"{v:.2f}%" if v is not None else "-"
    def fi(v): return f"{v:.2f}배" if v is not None else "-"
    def fe(v): return f"{v:,}원"   if v is not None else "-"

    # Sheet1~3: 회사별
    for corp in corps:
        ws = wb.create_sheet(corp["corp_name"][:28])
        ws.column_dimensions["A"].width = 22
        for col in ["B","C","D","E"]: ws.column_dimensions[col].width = 14
        ws.merge_cells("A1:E1")
        t = ws["A1"]; t.value = f"{corp['corp_name']}  3개년 재무제표"
        t.font = Font(name="맑은 고딕",size=13,bold=True,color="FFFFFF")
        t.fill = fill("1F3864"); t.alignment = Alignment(horizontal="center",vertical="center")
        ws.row_dimensions[1].height = 28
        for ci,h in enumerate(["항목","2023","2024","2025","YoY(24→25)"],1):
            c = ws.cell(row=2,column=ci,value=h)
            c.font=Font(name="맑은 고딕",bold=True,color="FFFFFF"); c.fill=fill("2F5496")
            c.alignment=Alignment(horizontal="center",vertical="center"); c.border=brd
        ROWS = [
            ("──손익계산서",None,"sec"),("매출(영업수익)","매출","amt"),
            ("매출원가","매출원가","amt"),("매출총이익","매출총이익","amt"),
            ("판매비와관리비","판매비와관리비","amt"),("영업이익","영업이익","amt"),
            ("금융수익","금융수익","amt"),("금융비용","금융비용","amt"),
            ("법인세비용차감전순이익","법인세비용차감전순이익","amt"),
            ("법인세비용","법인세비용","amt"),("당기순이익","당기순이익","amt"),
            ("EPS","EPS","eps"),
            ("──재무상태표",None,"sec"),("유동자산","유동자산","amt"),
            ("비유동자산","비유동자산","amt"),("자산총계","자산총계","amt"),
            ("유동부채","유동부채","amt"),("비유동부채","비유동부채","amt"),
            ("부채총계","부채총계","amt"),("이익잉여금","이익잉여금","amt"),
            ("자본총계","자본총계","amt"),
            ("──투자지표",None,"sec"),("영업이익률","영업이익률","rat"),
            ("순이익률","순이익률","rat"),("ROE","ROE","rat"),
            ("ROA","ROA","rat"),("부채비율","부채비율","rat"),
            ("유동비율","유동비율","rat"),("이자보상배율","이자보상배율","idx"),
        ]
        ri = 3
        for label,key,kind in ROWS:
            if kind=="sec":
                ws.merge_cells(start_row=ri,start_column=1,end_row=ri,end_column=5)
                c=ws.cell(row=ri,column=1,value=label)
                c.font=Font(name="맑은 고딕",bold=True,color="1F3864")
                c.fill=fill("D6E0F5"); c.border=brd; ri+=1; continue
            vals = {yr: corp["data"][str(yr)].get(key) for yr in YEARS}
            fmt = {"amt":fa,"rat":fp,"idx":fi,"eps":fe}.get(kind,fa)
            c=ws.cell(row=ri,column=1,value=label)
            c.font=Font(name="맑은 고딕"); c.fill=fill("F5F8FF"); c.border=brd
            for ci,yr in enumerate(YEARS,2):
                cell=ws.cell(row=ri,column=ci,value=fmt(vals[yr]))
                cell.font=Font(name="맑은 고딕"); cell.fill=fill("F5F8FF")
                cell.alignment=Alignment(horizontal="right"); cell.border=brd
            v25,v24=vals[2025],vals[2024]
            yoy="-"
            if v25 is not None and v24 and v24!=0:
                r=(v25-v24)/abs(v24)*100; yoy=f"{'▲' if r>=0 else '▼'}{abs(r):.1f}%"
            yc=ws.cell(row=ri,column=5,value=yoy)
            yc.alignment=Alignment(horizontal="center"); yc.border=brd
            yc.font=Font(name="맑은 고딕",color=("006400" if yoy.startswith("▲") else ("8B0000" if yoy.startswith("▼") else "000000")),bold=True)
            ri+=1

    # Sheet4: 비교
    ws4=wb.create_sheet("전체비교(2025)")
    ws4.column_dimensions["A"].width=22
    for i in range(3): ws4.column_dimensions[chr(66+i)].width=16
    ws4.merge_cells("A1:D1")
    t=ws4["A1"]; t.value="  ·  ".join(c["corp_name"] for c in corps)+" 비교 (2025)"
    t.font=Font(name="맑은 고딕",size=12,bold=True,color="FFFFFF")
    t.fill=fill("1F3864"); t.alignment=Alignment(horizontal="center",vertical="center")
    ws4.row_dimensions[1].height=26
    for ci,(h,col) in enumerate(zip(["항목"]+[c["corp_name"] for c in corps],["2F5496","1F3864","2E75B6","375623"]),1):
        c=ws4.cell(row=2,column=ci,value=h)
        c.font=Font(name="맑은 고딕",bold=True,color="FFFFFF"); c.fill=fill(col)
        c.alignment=Alignment(horizontal="center"); c.border=brd
    ri=3
    for label,key,kind in ROWS:
        if kind=="sec":
            ws4.merge_cells(start_row=ri,start_column=1,end_row=ri,end_column=4)
            c=ws4.cell(row=ri,column=1,value=label)
            c.font=Font(name="맑은 고딕",bold=True,color="1F3864"); c.fill=fill("D6E0F5"); c.border=brd; ri+=1; continue
        fmt={"amt":fa,"rat":fp,"idx":fi,"eps":fe}.get(kind,fa)
        c=ws4.cell(row=ri,column=1,value=label)
        c.font=Font(name="맑은 고딕"); c.fill=fill("F5F8FF"); c.border=brd
        for ci,corp in enumerate(corps,2):
            v=corp["data"]["2025"].get(key)
            cell=ws4.cell(row=ri,column=ci,value=fmt(v))
            cell.font=Font(name="맑은 고딕"); cell.fill=fill("F5F8FF")
            cell.alignment=Alignment(horizontal="right"); cell.border=brd
        ri+=1

    if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
    wb.save(path)
